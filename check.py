# -----------------------------------------------------
# check.py : script python per analizzare un file .xlsx
# -----------------------------------------------------
#
# example:
#
#  py .\check.py .\HT123456_Report_TPS_Ballò_20251114.xlsx
#
# --------------------------------------------------------

from openpyxl import load_workbook
import sys,argparse

# Creiamo il parser dei parametri da linea di comando
# parser = argparse.ArgumentParser(description="Fa il check un file xlsx elencando gli sheets e le tabelle presenti in esso.")
# parser.add_argument("xlsxfile", help="Percorso del file .xlsx")

# Leggiamo i parametri
# args = parser.parse_args()
# xlsx_path = args.xlsxfile

def list_sheets_and_tables(xlsx_file):
    # Carica il workbook
    wb = load_workbook(xlsx_file, data_only=True)
    
    # Itera sui fogli
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Foglio: {sheet_name}")
        
        # Verifica se ci sono tabelle embedded
        if ws._tables:
            try:
                for table in ws._tables:
                    print(f"  Tabella: {table.name}")
            except AttributeError:
                print(f" La tabella {table} non ha un nome")                        
        else:
            print("  Nessuna tabella presente")

    wb.close()

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python check.py <file.xlsx>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    list_sheets_and_tables(file_path)
