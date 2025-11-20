# -------------------------------------------------------------------------------------------
#
# xlsx_files_diff.py :  visualizza le differenze fra due files excel .xlsx
#                
# -------------------------------------------------------------------------------------------
#
# written by Massimiliano Cosmelli ( @_°° massimiliano.cosmelli@accelleron-industries.com )
#
#                   CopyRight 2025-2026 Accelleron Industries 
#
# -------------------------------------------------------------------------------------------

import openpyxl
import os,sys
from pathlib import Path

def check_diff(file1,file2):
    
    file1 = Path(f"{file1}")
    file2 = Path(f"{file2}") 

    wb1 = openpyxl.load_workbook(file1, data_only=True)
    wb2 = openpyxl.load_workbook(file2, data_only=True)

    for ws1, ws2 in zip(wb1.worksheets, wb2.worksheets):
        for row1, row2 in zip(ws1.iter_rows(values_only=True), ws2.iter_rows(values_only=True)):
            if row1 != row2:
                print("Differenze trovate:", row1, row2)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Uso: file_identity_test.py <file1> <file2>")
        sys.exit(1)
        
    check_diff(sys.argv[1],sys.argv[2])