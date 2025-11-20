# -----------------------------------------------------
# check.py : script python per analizzare un file .xlsx
# -----------------------------------------------------
#
# example:
#
#  py .\check.py .\HT123456_Report_TPS_Ballò_20251114.xlsx
#
# --------------------------------------------------------
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import sys,os,zipfile

REQUIRED_FILES = [
    "xl/workbook.xml",
    "xl/sharedStrings.xml",
    "xl/styles.xml",
]

# Funzione cross-platform con openpyxl
def check_openpyxl(file_path):
    try:
        wb = load_workbook(file_path, data_only=True)
        wb.close()
        print(f"[openpyxl] {file_path} aperto senza errori.")
        return True
    except Exception as e:
        print(f"[openpyxl] Errore nell'aprire {file_path}: {e}")
        return False

# Funzione solo Windows con Excel COM
def check_excel_com(file_path):
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.DisplayAlerts = False  # evita popup
        file_path = os.path.abspath(file_path)
        wb = excel.Workbooks.Open(file_path)
        wb.Close(False)
        excel.Quit()
        print(f"[Excel COM] {file_path} aperto senza errori di recovery.")
        return True
    except ImportError:
        print("[Excel COM] pywin32 non installato, salto il controllo COM.")
        return None
    except Exception as e:
        print(f"[Excel COM] Errore nell'aprire {file_path} con Excel: {e}")
        return False
        
def list_sheets_and_tables(xlsx_file):
    wb = load_workbook(xlsx_file, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Foglio: {sheet_name}")
        
        if ws._tables:
            for table in ws._tables:
                # ws._tables può contenere oggetti Table o semplici stringhe
                if hasattr(table, "name"):
                    print(f"  Tabella: {table.name}")
                else:
                    print(f"  Tabella: {table}")
        else:
            print("  Nessuna tabella presente")
    wb.close()

def check_xlsx_detailed(file_path):
    if not os.path.exists(file_path):
        print(f"File non trovato: {file_path}")
        return False

    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            namelist = z.namelist()
            ok = True

            # Controllo file obbligatori
            missing_files = [f for f in REQUIRED_FILES if f not in namelist]
            if missing_files:
                print("❌ File obbligatori mancanti:", missing_files)
                ok = False
            else:
                print("✅ File obbligatori presenti")

            # Controllo fogli
            worksheets = [f for f in namelist if f.startswith("xl/worksheets/")]
            if not worksheets:
                print("❌ Nessun foglio trovato in xl/worksheets/")
                ok = False
            else:
                print("📄 Fogli trovati:")
                for f in worksheets:
                    print(f"  - {f}")

            # Controllo tabelle
            tables = [f for f in namelist if f.startswith("xl/tables/")]
            if tables:
                print("📊 Tabelle trovate:")
                for f in tables:
                    print(f"  - {f}")
            else:
                print("ℹ️ Nessuna tabella trovata")

            # Controllo XML interno
            print("🔍 Controllo XML interni...")
            for f in [f for f in namelist if f.endswith(".xml")]:
                try:
                    with z.open(f) as xml_file:
                        ET.parse(xml_file)
                except ET.ParseError as e:
                    print(f"❌ XML malformato: {f} --> {e}")
                    ok = False

        if ok:
            print("✅ Struttura XLSX OK (tutti XML ben formati)")
        else:
            print("⚠️ Problemi rilevati nella struttura XLSX")
        return ok

    except zipfile.BadZipFile:
        print("❌ File non è un XLSX valido (zip malformato)")
        return False
        
if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python check.py <file.xlsx>")
        sys.exit(1)

    file_path = sys.argv[1]
    # list_sheets_and_tables(file_path)
    # quit()

    print("=== Controllo openpyxl ===")
    openpyxl_ok = check_openpyxl(file_path)

    print("\n=== Controllo Excel COM (Windows) ===")
    excel_ok = check_excel_com(file_path)

    print("\n=== Risultato finale ===")
    if openpyxl_ok and (excel_ok is True or excel_ok is None):
        print("File OK: leggibile e (se Windows) Excel lo apre senza errori.")
        # effettua ulteriori test
        list_sheets_and_tables(file_path)
        check_xlsx_detailed(file_path)
    elif openpyxl_ok and excel_ok is False:
        print("File leggibile, ma Excel segnala problemi di recovery!")
        # effettua ulteriori test
        list_sheets_and_tables(file_path)
        check_xlsx_detailed(file_path)
    else:
        print("File corrotto o non leggibile.")