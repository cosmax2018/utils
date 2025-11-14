import zipfile
import xml.etree.ElementTree as ET
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

REQUIRED_FILES = [
    "xl/workbook.xml",
    "xl/sharedStrings.xml",
    "xl/styles.xml",
]

def check_xlsx_structure(file_path):
    """Controllo struttura interna come zip/XML"""
    print("\n=== Controllo struttura XLSX (ZIP/XML) ===")
    if not os.path.exists(file_path):
        print(f"❌ File non trovato: {file_path}")
        return False

    ok = True
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            namelist = z.namelist()

            # File obbligatori
            missing_files = [f for f in REQUIRED_FILES if f not in namelist]
            if missing_files:
                print("❌ File obbligatori mancanti:", missing_files)
                ok = False
            else:
                print("✅ File obbligatori presenti")

            # Fogli
            worksheets = [f for f in namelist if f.startswith("xl/worksheets/")]
            if worksheets:
                print("📄 Fogli trovati:")
                for f in worksheets:
                    print(f"  - {f}")
            else:
                print("❌ Nessun foglio trovato")
                ok = False

            # Tabelle
            tables = [f for f in namelist if f.startswith("xl/tables/")]
            if tables:
                print("📊 Tabelle trovate:")
                for f in tables:
                    print(f"  - {f}")
            else:
                print("ℹ️ Nessuna tabella trovata")

            # XML interni
            print("🔍 Controllo XML interni...")
            for f in [f for f in namelist if f.endswith(".xml")]:
                try:
                    with z.open(f) as xml_file:
                        ET.parse(xml_file)
                except ET.ParseError as e:
                    print(f"❌ XML malformato: {f} --> {e}")
                    ok = False

        if ok:
            print("✅ Struttura XLSX OK")
        else:
            print("⚠️ Problemi rilevati nella struttura XLSX")
        return ok
    except zipfile.BadZipFile:
        print("❌ File non è un XLSX valido (zip malformato)")
        return False

def check_openpyxl(file_path):
    """Elenca fogli e tabelle con openpyxl"""
    print("\n=== Controllo fogli e tabelle con openpyxl ===")
    if load_workbook is None:
        print("⚠️ openpyxl non installato, salto questo controllo")
        return

    try:
        wb = load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"📄 Foglio: {sheet_name}")
            if ws._tables:
                for t in ws._tables:
                    name = getattr(t, "name", t)
                    print(f"  📊 Tabella: {name}")
            else:
                print("  ℹ️ Nessuna tabella")
        wb.close()
    except Exception as e:
        print(f"❌ Errore openpyxl: {e}")

def check_excel_com(file_path):
    """Verifica reale con Excel COM (solo Windows + Office)"""
    print("\n=== Controllo Excel COM (Windows) ===")
    try:
        import win32com.client
    except ImportError:
        print("⚠️ pywin32 non installato, salto il controllo COM")
        return

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(file_path)
        wb.Close(False)
        excel.Quit()
        print("✅ Excel COM: file aperto senza errori di recovery")
    except Exception as e:
        print(f"❌ Excel COM: errore nell'aprire il file: {e}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python check_xlsx_full.py <file.xlsx>")
        sys.exit(1)

    file_path = os.path.abspath(sys.argv[1])

    # 1. Controllo struttura
    check_xlsx_structure(file_path)

    # 2. Openpyxl
    check_openpyxl(file_path)

    # 3. Excel COM (solo Windows)
    check_excel_com(file_path)
