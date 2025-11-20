import openpyxl
import sys
import os

def keep_only_sheet(input_path, sheet_to_keep, output_path=None):
    """
    Mantiene solo il foglio sheet_to_keep in un file .xlsx esistente.
    Tutti gli altri fogli vengono eliminati senza rompere sharedStrings o tabelle.
    """
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_filtered.xlsx"

    # Apri il workbook esistente
    wb = openpyxl.load_workbook(input_path)

    if sheet_to_keep not in wb.sheetnames:
        print(f"❌ Errore: il foglio '{sheet_to_keep}' non esiste nel workbook.")
        return

    # Rimuovi tutti i fogli tranne sheet_to_keep
    for sheetname in wb.sheetnames:
        if sheetname != sheet_to_keep:
            std = wb[sheetname]
            wb.remove(std)

    # Salva il file modificato
    wb.save(output_path)
    print(f"✅ File salvato mantenendo solo il foglio '{sheet_to_keep}': {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Uso: python keep_only_sheet.py <file.xlsx> <foglio_da_mantenere>")
        sys.exit(1)

    file_path = os.path.abspath(sys.argv[1])
    sheet_name = sys.argv[2]

    keep_only_sheet(file_path, sheet_name)
