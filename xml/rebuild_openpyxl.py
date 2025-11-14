import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment
import sys
import os

def copy_cell(source_cell, target_cell):
    """Copia valore, formula e stili base di una cella"""
    target_cell.value = source_cell.value
    target_cell.data_type = source_cell.data_type
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.alignment = source_cell.alignment.copy()
        target_cell.number_format = source_cell.number_format

def sanitize_name(name):
    """Rende il nome compatibile con Excel: ASCII, max 31 caratteri, caratteri validi"""
    import re
    name = re.sub(r'[\\/*?:[\]]', '_', name)
    name = ''.join(c if ord(c) < 128 else '_' for c in name)
    return name[:31]

def rebuild_workbook(input_path, output_path=None):
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_clean.xlsx"

    # Apri workbook sorgente
    wb_src = openpyxl.load_workbook(input_path, data_only=False, keep_links=False)

    # Crea nuovo workbook
    wb_new = openpyxl.Workbook()
    # Rimuovi foglio di default creato da openpyxl
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)

    for ws_src in wb_src.worksheets:
        # Crea nuovo foglio
        name_safe = sanitize_name(ws_src.title)
        ws_new = wb_new.create_sheet(title=name_safe)

        # Copia celle con valori, formule e stili base
        for row in ws_src.iter_rows():
            for cell in row:
                target_cell = ws_new.cell(row=cell.row, column=cell.column)
                copy_cell(cell, target_cell)

        # Copia dimensioni colonne e righe
        for col_letter, col_dim in ws_src.column_dimensions.items():
            if col_letter in ws_new.column_dimensions:
                ws_new.column_dimensions[col_letter].width = col_dim.width
        for row_idx, row_dim in ws_src.row_dimensions.items():
            ws_new.row_dimensions[row_idx].height = row_dim.height

    # Salva file pulito
    wb_new.save(output_path)
    print(f"✅ File XLSX ricreato e pulito: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python rebuild_openpyxl.py <file.xlsx>")
        sys.exit(1)

    file_path = os.path.abspath(sys.argv[1])
    rebuild_workbook(file_path)
