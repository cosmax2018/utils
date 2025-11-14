import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import sys
import os

from copy import copy  # aggiungi in cima al file

def copy_cell(source_cell, target_cell):
    """Copia valore e stili base di una cella, senza forzare data_type"""
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format



def sanitize_name(name):
    """Rende il nome compatibile con Excel: ASCII, max 31 caratteri, caratteri validi"""
    import re
    name = re.sub(r'[\\/*?:[\]]', '_', name)
    name = ''.join(c if ord(c) < 128 else '_' for c in name)
    return name[:31]

def rebuild_workbook_with_tables(input_path, output_path=None):
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_clean_tables.xlsx"

    # Apri workbook sorgente
    wb_src = openpyxl.load_workbook(input_path, data_only=False, keep_links=False)

    # Crea nuovo workbook
    wb_new = openpyxl.Workbook()
    # Rimuovi foglio di default creato da openpyxl
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)

    for ws_src in wb_src.worksheets:
        # Crea nuovo foglio
        name_safe = sanitize_name(ws_src.title)
        ws_new = wb_new.create_sheet(title=name_safe)

        # Copia celle con valori, formule e stili base
        for row in ws_src.iter_rows():
            for cell in row:
                target_cell = ws_new.cell(row=cell.row, column=cell.column)
                copy_cell(cell, target_cell)

        # Copia dimensioni colonne e righe
        for col_letter, col_dim in ws_src.column_dimensions.items():
            if col_letter in ws_new.column_dimensions:
                ws_new.column_dimensions[col_letter].width = col_dim.width
        for row_idx, row_dim in ws_src.row_dimensions.items():
            ws_new.row_dimensions[row_idx].height = row_dim.height

        # Copia tabelle se il range è valido
        for tbl in ws_src._tables:
            # Verifica che il range esista
            try:
                ws_new[tbl.ref]
            except:
                continue  # Range non valido, salta
            # Crea nuova tabella
            new_table = Table(displayName=sanitize_name(tbl.displayName), ref=tbl.ref)
            # Copia stile se presente
            if tbl.tableStyleInfo:
                tsi = TableStyleInfo(name=tbl.tableStyleInfo.name,
                                     showFirstColumn=tbl.tableStyleInfo.showFirstColumn,
                                     showLastColumn=tbl.tableStyleInfo.showLastColumn,
                                     showRowStripes=tbl.tableStyleInfo.showRowStripes,
                                     showColumnStripes=tbl.tableStyleInfo.showColumnStripes)
                new_table.tableStyleInfo = tsi
            ws_new.add_table(new_table)

    # Salva file pulito con tabelle
    wb_new.save(output_path)
    print(f"✅ File XLSX ricreato e pulito con tabelle valide: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python rebuild_with_tables.py <file.xlsx>")
        sys.exit(1)

    file_path = os.path.abspath(sys.argv[1])
    rebuild_workbook_with_tables(file_path)
