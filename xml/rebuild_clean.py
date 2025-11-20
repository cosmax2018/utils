import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import sys
import os
import copy

def sanitize_name(name):
    """Rende il nome compatibile con Excel: ASCII, max 31 caratteri"""
    import re
    name = re.sub(r'[\\/*?:[\]]', '_', name)
    name = ''.join(c if ord(c) < 128 else '_' for c in name)
    return name[:31]

def copy_cell(source_cell, target_cell):
    """Copia valore, formula e stili base di una cella"""
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format

def copy_table(ws_src, ws_target):
    """Copia tabelle dal foglio sorgente a quello target se il range è valido"""
    for tbl in ws_src._tables:
        try:
            ws_target[tbl.ref]  # verifica che il range esista
        except:
            continue
        new_table = Table(displayName=sanitize_name(tbl.displayName), ref=tbl.ref)
        if tbl.tableStyleInfo:
            tsi = TableStyleInfo(name=tbl.tableStyleInfo.name,
                                 showFirstColumn=tbl.tableStyleInfo.showFirstColumn,
                                 showLastColumn=tbl.tableStyleInfo.showLastColumn,
                                 showRowStripes=tbl.tableStyleInfo.showRowStripes,
                                 showColumnStripes=tbl.tableStyleInfo.showColumnStripes)
            new_table.tableStyleInfo = tsi
        ws_target.add_table(new_table)

def rebuild_workbook(input_path, sheets_to_keep, output_path=None):
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_clean.xlsx"

    # Apri workbook sorgente
    wb_src = openpyxl.load_workbook(input_path, data_only=False, keep_links=False)

    # Crea nuovo workbook
    wb_new = openpyxl.Workbook()
    # Rimuovi foglio di default creato da openpyxl
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)

    for sheet_name in sheets_to_keep:
        if sheet_name not in wb_src.sheetnames:
            print(f"❌ Foglio '{sheet_name}' non trovato nel workbook sorgente.")
            continue

        ws_src = wb_src[sheet_name]
        ws_new = wb_new.create_sheet(title=sanitize_name(ws_src.title))

        # Copia celle con valori, formule e stili
        for row in ws_src.iter_rows():
            for cell in row:
                target_cell = ws_new.cell(row=cell.row, column=cell.column)
                copy_cell(cell, target_cell)

        # Copia dimensioni colonne e righe
        for col_letter, col_dim in ws_src.column_dimensions.items():
            ws_new.column_dimensions[col_letter].width = col_dim.width
        for row_idx, row_dim in ws_src.row_dimensions.items():
            ws_new.row_dimensions[row_idx].height = row_dim.height

        # Copia tabelle
        copy_table(ws_src, ws_new)

    # Copia nomi definiti validi che puntano solo ai fogli tenuti
    if hasattr(wb_src.defined_names, "definedName") and wb_src.defined_names.definedName:
        for dn in wb_src.defined_names.definedName:
            # Verifica se il foglio esiste ancora
            valid = False
            for sheet_name in sheets_to_keep:
                if sheet_name in dn.attr_text:
                    valid = True
                    break
            if valid:
                wb_new.defined_names.append(dn)

    # Salva workbook pulito
    wb_new.save(output_path)
    print(f"✅ File ricreato correttamente: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python rebuild_clean.py <file.xlsx> <foglio1> [<foglio2> ...]")
        sys.exit(1)

    file_path = os.path.abspath(sys.argv[1])
    sheets = sys.argv[2:]
    rebuild_workbook(file_path, sheets)
